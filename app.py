import streamlit as st
import pandas as pd
import re
import os
import numpy as np
from io import BytesIO
from openai import OpenAI
client = OpenAI()
#  VIC FANGIO LLM 
SYSTEM_PROMPT = """
You are a defensive coordinator modeled after Vic Fangio.

Background:
- Legendary NFL defensive coordinator known for disguising coverages and confusing quarterbacks.
- Built defenses around pattern-matching zone concepts that look like man coverage pre-snap.
- Emphasizes stopping the run first, controlling the line of scrimmage, and making quarterbacks uncomfortable.
- Known for meticulous preparation and exploiting offensive tendencies.

Style:
- Direct, no-nonsense, and detail-oriented.
- Use specific terminology: cover 2, quarters, pattern-match, two-high shell, box count, leverage.
- Focus on taking away the opponents best play and forcing them into uncomfortable situations.

When a coach asks a question:
1. Identify the offensive tendency most relevant to the question.
2. Give 3-6 concrete defensive adjustments or game-plan points in bullets.
3. Keep it concise enough to scan quickly before a game.
"""
def summarize_view(df: pd.DataFrame) -> str:
    if df.empty:
        return "No plays in view."
    total = len(df)
    run_rate = round(df[cols['type']].eq("RUN").mean() * 100, 1)
    pass_rate = round(df[cols['type']].eq("PASS").mean() * 100, 1)
    avg_gain  = round(df[cols['gain']].mean(), 2)
    top_forms = df[cols['form']].value_counts().head(3).to_dict()
    top_pers  = df["PERSONNEL"].value_counts().head(3).to_dict()
    lines = [
        f"Total plays in view: {total}.",
        f"Run rate: {run_rate}pct, Pass rate: {pass_rate}pct.",
        f"Average gain: {avg_gain} yards.",
        f"Top formations: {top_forms}.",
        f"Top personnel groups: {top_pers}.",
    ]
    return "\n".join(lines)

def call_fangio_llm(summary_text: str, question: str) -> str:
    resp = client.responses.create(
        model="gpt-4.1-mini",
        input=[
            {"role": "system", "content": FANGIO_SYSTEM_PROMPT},
            {"role": "user",
             "content": f"Opponent tendency summary:\n{summary_text}\n\nCoach question:\n{question}"}
        ],
    )
    return resp.output[0].content[0].text


# ============================================================
# EXCEL EXPORT
# ============================================================

def build_excel_export(export_dict, p_data, drive_dla, pers_dla,
                       fpar_df, sss_summary,
                       sss_by_form, chain, intel_df, scout_sections,
                       cols, verdict_score):

    from openpyxl import Workbook
    from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = Workbook()
    wb.remove(wb.active)

    #  COLOR PALETTE 
    RED    = "C0392B"
    YELLOW = "F39C12"
    GREEN  = "27AE60"
    DARK   = "1C2833"
    MED    = "2E4057"
    LIGHT  = "D6EAF8"
    WHITE  = "FFFFFF"
    GRAY   = "F2F3F4"
    BORDER = "BDC3C7"

    grade_colors = {
        'A': ('1E8449', WHITE),
        'B': ('27AE60', WHITE),
        'C': ('F39C12', WHITE),
        'D': ('E67E22', WHITE),
        'F': ('C0392B', WHITE),
    }

    def make_fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def make_font(bold=False, color=WHITE, size=11):
        return Font(bold=bold, color=color, size=size, name='Calibri')

    def make_border():
        s = Side(style='thin', color=BORDER)
        return Border(left=s, right=s, top=s, bottom=s)

    def make_align(wrap=False, h='left', v='center'):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    def set_col_width(ws, col, width):
        ws.column_dimensions[get_column_letter(col)].width = width

    def header_row(ws, row_num, values, bg=DARK, fg=WHITE, bold=True, sizes=None):
        for i, val in enumerate(values, 1):
            c = ws.cell(row=row_num, column=i, value=val)
            c.fill    = make_fill(bg)
            c.font    = Font(bold=bold, color=fg, size=sizes[i-1] if sizes else 11, name='Calibri')
            c.alignment = make_align(h='center')
            c.border  = make_border()

    def data_row(ws, row_num, values, bg=WHITE, fg='000000', bold=False, wrap=False):
        for i, val in enumerate(values, 1):
            c = ws.cell(row=row_num, column=i, value=val)
            c.fill      = make_fill(bg)
            c.font      = Font(bold=bold, color=fg, size=10, name='Calibri')
            c.alignment = make_align(wrap=wrap)
            c.border    = make_border()

    def section_title(ws, row_num, title, ncols, bg=MED):
        ws.merge_cells(start_row=row_num, start_column=1,
                       end_row=row_num, end_column=ncols)
        c = ws.cell(row=row_num, column=1, value=title)
        c.fill      = make_fill(bg)
        c.font      = Font(bold=True, color=WHITE, size=12, name='Calibri')
        c.alignment = make_align(h='center')

    def grade_cell(ws, row, col, grade):
        bg, fg = grade_colors.get(str(grade), ('FFFFFF', '000000'))
        c = ws.cell(row=row, column=col, value=grade)
        c.fill      = make_fill(bg)
        c.font      = Font(bold=True, color=fg, size=10, name='Calibri')
        c.alignment = make_align(h='center')
        c.border    = make_border()

    total    = len(p_data)
    runs     = (p_data[cols['type']] == 'RUN').sum()
    passes   = (p_data[cols['type']] == 'PASS').sum()
    run_pct  = round(runs / total * 100) if total else 0
    pass_pct = round(passes / total * 100) if total else 0
    avg_gain = round(p_data[cols['gain']].mean(), 1)
    fd_rate  = round(p_data['Is_FD'].mean() * 100)
    succ_rt  = round(p_data['Is_Succ'].mean() * 100)
    exp_rt   = round(p_data['Is_Explosive'].mean() * 100)

    if verdict_score >= 7:
        verdict_text  = "HIGH THREAT"
        verdict_color = RED
    elif verdict_score >= 4:
        verdict_text  = "MODERATE THREAT"
        verdict_color = YELLOW
    else:
        verdict_text  = "MANAGEABLE"
        verdict_color = GREEN

    # 
    # SHEET 1  EXECUTIVE SUMMARY
    # 
    ws1 = wb.create_sheet("1 - Executive Summary")
    ws1.sheet_view.showGridLines = False

    # Title banner
    ws1.merge_cells("A1:H1")
    c = ws1["A1"]
    c.value     = "FormationIQ  Opponent Scouting Report"
    c.fill      = make_fill(DARK)
    c.font      = Font(bold=True, color=WHITE, size=18, name='Calibri')
    c.alignment = make_align(h='center')
    ws1.row_dimensions[1].height = 36

    # Verdict banner
    ws1.merge_cells("A2:H2")
    c = ws1["A2"]
    c.value     = f"SCOUTING VERDICT: {verdict_text}"
    c.fill      = make_fill(verdict_color)
    c.font      = Font(bold=True, color=WHITE, size=14, name='Calibri')
    c.alignment = make_align(h='center')
    ws1.row_dimensions[2].height = 28

    # Key stats header
    r = 4
    section_title(ws1, r, "KEY OFFENSIVE METRICS", 8)
    ws1.row_dimensions[r].height = 22

    r = 5
    header_row(ws1, r,
               ["Total Plays","Run Plays","Pass Plays","Run %","Pass %",
                "Avg Gain","FD Rate","Success Rate"])
    r = 6
    data_row(ws1, r,
             [total, int(runs), int(passes), f"{run_pct}pct", f"{pass_pct}pct",
              f"{avg_gain} yds", f"{fd_rate}pct", f"{succ_rt}pct"],
             bg=LIGHT, fg='000000', bold=True)

    for col in range(1, 9):
        set_col_width(ws1, col, 14)

    # Top formations
    r = 8
    section_title(ws1, r, "TOP FORMATIONS BY VOLUME", 4)
    r = 9
    header_row(ws1, r, ["Formation","Plays","Avg Gain","FD Rate %"], bg=MED)
    top_forms = (
        p_data.groupby(cols['form'])
        .agg(Plays=(cols['gain'],'count'), Avg_Gain=(cols['gain'],'mean'), FD_Rate=('Is_FD','mean'))
        .sort_values('Plays', ascending=False).head(5)
    )
    for i, (form, row) in enumerate(top_forms.iterrows()):
        r += 1
        bg = GRAY if i % 2 == 0 else WHITE
        data_row(ws1, r, [form, int(row['Plays']),
                          round(row['Avg_Gain'], 1),
                          f"{round(row['FD_Rate']*100)}pct"], bg=bg, fg='000000')

    set_col_width(ws1, 1, 22)
    set_col_width(ws1, 2, 10)
    set_col_width(ws1, 3, 12)
    set_col_width(ws1, 4, 12)

    # Top chain movers
    r += 2
    section_title(ws1, r, "TOP CHAIN-MOVING PLAYS", 4)
    r += 1
    header_row(ws1, r, ["Play","Plays","FD Rate %","Success Rate %"], bg=MED)
    for i, (play, row) in enumerate(chain.head(5).iterrows()):
        r += 1
        bg = GRAY if i % 2 == 0 else WHITE
        data_row(ws1, r, [play, int(row['Plays']),
                          f"{row['FD Rate %']}pct",
                          f"{row['Success Rate %']}pct"], bg=bg, fg='000000')

    # Top exploits
    r += 2
    section_title(ws1, r, "TOP DEFENSIVE EXPLOITS", 5)
    r += 1
    header_row(ws1, r, ["Category","Finding","Action","Stat","Priority"], bg=MED)
    exploits = []
    if not sss_summary.empty:
        top_sss = sss_summary.sort_values('Stress %', ascending=False).iloc[0]
        exploits.append([
            "SSS",
            f"{top_sss.name} creates {int(top_sss['Stress %'])}pct of stress situations",
            f"Stop their {top_sss.name.lower()} on early downs",
            f"Avg prior gain: {round(top_sss['Avg_Prior_Gain'],1)} yds",
            "HIGH"
        ])
    if not intel_df.empty:
        for _, row in intel_df.head(3).iterrows():
            exploits.append([
                row['Category'],
                row['Signal'],
                row['Coaching Note'],
                row['Stat'],
                "MED"
            ])
    for i, exp in enumerate(exploits[:6]):
        r += 1
        bg = GRAY if i % 2 == 0 else WHITE
        data_row(ws1, r, exp, bg=bg, fg='000000', wrap=True)
        ws1.row_dimensions[r].height = 40

    set_col_width(ws1, 1, 12)
    set_col_width(ws1, 2, 35)
    set_col_width(ws1, 3, 45)
    set_col_width(ws1, 4, 16)
    set_col_width(ws1, 5, 10)

    # 
    # SHEET 2  PERSONNEL ANALYSIS
    # 
    ws4 = wb.create_sheet("4 - Personnel Analysis")
    ws4.sheet_view.showGridLines = False

    ws4.merge_cells("A1:J1")
    c = ws4["A1"]
    c.value     = "Personnel Group Analysis  Drive Leverage, Efficiency & Tendencies"
    c.fill      = make_fill(DARK)
    c.font      = Font(bold=True, color=WHITE, size=14, name='Calibri')
    c.alignment = make_align(h='center')
    ws4.row_dimensions[1].height = 28

    if not pers_dla.empty:
        r = 3
        section_title(ws4, r, "PERSONNEL LEVERAGE PROFILE", 10, bg=MED)
        r += 1
        header_row(ws4, r,
                   ["Personnel","Plays","DLS","Grade","Avg Gain",
                    "FD Rate %","Success %","Run %","Pass %","Explosive %"],
                   bg=DARK)
        for i, (pers, row) in enumerate(pers_dla.sort_values('Plays', ascending=False).iterrows()):
            r += 1
            bg = GRAY if i % 2 == 0 else WHITE
            data_row(ws4, r,
                     [pers, int(row['Plays']), row['DLS'], '',
                      row['Avg_Gain'], f"{row['FD_Rate']}pct",
                      f"{row['Success_Rate']}pct", f"{row['Run%']}pct",
                      f"{row['Pass%']}pct", f"{row['Explosive_Rt']}pct"],
                     bg=bg, fg='000000')
            grade_cell(ws4, r, 4, row['DLS_Grade'])

    for col, w in zip(range(1, 11), [12, 8, 8, 8, 10, 10, 10, 8, 8, 12]):
        set_col_width(ws4, col, w)

    # 
    # SHEET 3  SITUATIONAL BREAKDOWNS
    # 
    ws5 = wb.create_sheet("5 - Situational Breakdowns")
    ws5.sheet_view.showGridLines = False

    ws5.merge_cells("A1:F1")
    c = ws5["A1"]
    c.value     = "Situational Analysis  Third Down, Red Zone & Field Position"
    c.fill      = make_fill(DARK)
    c.font      = Font(bold=True, color=WHITE, size=14, name='Calibri')
    c.alignment = make_align(h='center')
    ws5.row_dimensions[1].height = 28

    # Third down
    t3 = p_data[p_data[cols['dn']] == 3].copy()
    if not t3.empty:
        t3['Sit'] = t3[cols['dist']].apply(
            lambda x: "Third & Short (1-3)" if x <= 3 else
                      ("Third & Mid (4-7)" if x <= 7 else "Third & Long (7+)")
        )
        r = 3
        section_title(ws5, r, "3RD DOWN EFFICIENCY", 5, bg=MED)
        r += 1
        header_row(ws5, r,
                   ["Situation","Plays","Conv Rate %","Run %","Pass %"],
                   bg=DARK)
        for sit in ["Third & Short (1-3)", "Third & Mid (4-7)", "Third & Long (7+)"]:
            sub = t3[t3['Sit'] == sit]
            if not sub.empty:
                r += 1
                run_p  = round((sub[cols['type']] == 'RUN').mean() * 100)
                pass_p = round((sub[cols['type']] == 'PASS').mean() * 100)
                data_row(ws5, r,
                         [sit, len(sub),
                          f"{round(sub['Is_FD'].mean()*100)}pct",
                          f"{run_p}pct", f"{pass_p}pct"],
                         bg=GRAY, fg='000000')

    # FPAR
    if not fpar_df.empty:
        r = ws5.max_row + 2
        section_title(ws5, r, "FIELD POSITION AGGRESSION (FPAR)", 6, bg=MED)
        r += 1
        fpar_reset = fpar_df.reset_index()
        cols_fpar  = ['Field_Zone', cols['dn'], 'Plays', 'Pass_Rate', 'Success_Rate', 'Avg_Gain']
        header_row(ws5, r,
                   ["Zone","Down","Plays","Pass Rate %","Success Rate %","Avg Gain"],
                   bg=DARK)
        for i, row in fpar_reset.iterrows():
            r += 1
            bg = GRAY if i % 2 == 0 else WHITE
            data_row(ws5, r,
                     [row['Field_Zone'], int(row[cols['dn']]),
                      int(row['Plays']), f"{row['Pass_Rate']}pct",
                      f"{row['Success_Rate']}pct", row['Avg_Gain']],
                     bg=bg, fg='000000')

    for col, w in zip(range(1, 7), [28, 8, 8, 12, 14, 10]):
        set_col_width(ws5, col, w)

    # 
    # SHEET 4  AI SCOUTING INTELLIGENCE
    # 
    ws6 = wb.create_sheet("6 - AI Scouting Intelligence")
    ws6.sheet_view.showGridLines = False

    ws6.merge_cells("A1:D1")
    c = ws6["A1"]
    c.value     = "AI Scouting Intelligence  Auto-Detected Behavioral Patterns & Tendencies"
    c.fill      = make_fill(DARK)
    c.font      = Font(bold=True, color=WHITE, size=14, name='Calibri')
    c.alignment = make_align(h='center')
    ws6.row_dimensions[1].height = 28

    if not intel_df.empty:
        r = 3
        current_cat = None
        header_row(ws6, r,
                   ["Category","Signal","Stat","Coaching Note"],
                   bg=DARK)
        for i, row in intel_df.iterrows():
            r += 1
            if row['Category'] != current_cat:
                current_cat = row['Category']
                bg = LIGHT
            else:
                bg = WHITE
            data_row(ws6, r,
                     [row['Category'], row['Signal'],
                      row['Stat'], row['Coaching Note']],
                     bg=bg, fg='000000', wrap=True)
            ws6.row_dimensions[r].height = 50

    for col, w in zip(range(1, 5), [14, 30, 20, 60]):
        set_col_width(ws6, col, w)

    # 
    # SHEET 5  SCOUT REPORT
    # 
    ws7 = wb.create_sheet("7 - Scout Report")
    ws7.sheet_view.showGridLines = False

    ws7.merge_cells("A1:B1")
    c = ws7["A1"]
    c.value     = "FormationIQ  Full Scouting Report"
    c.fill      = make_fill(DARK)
    c.font      = Font(bold=True, color=WHITE, size=16, name='Calibri')
    c.alignment = make_align(h='center')
    ws7.row_dimensions[1].height = 32

    r = 2
    for section_title_text, section_body in scout_sections:
        r += 1
        ws7.merge_cells(f"A{r}:B{r}")
        c = ws7.cell(row=r, column=1, value=section_title_text)
        c.fill      = make_fill(MED)
        c.font      = Font(bold=True, color=WHITE, size=12, name='Calibri')
        c.alignment = make_align(h='left')
        ws7.row_dimensions[r].height = 24

        r += 1
        clean_body = re.sub(r'\*\*|__', '', str(section_body)).strip()
        ws7.merge_cells(f"A{r}:B{r}")
        c = ws7.cell(row=r, column=1, value=clean_body)
        c.fill      = make_fill(WHITE)
        c.font      = Font(color='000000', size=10, name='Calibri')
        c.alignment = make_align(wrap=True, h='left', v='top')
        line_count  = max(clean_body.count('\n') + 1, 3)
        ws7.row_dimensions[r].height = min(line_count * 15, 200)
        r += 1

    ws7.column_dimensions['A'].width = 40
    ws7.column_dimensions['B'].width = 80

    # 
    # SHEET 6  RAW PLAY-BY-PLAY
    # 
    ws8 = wb.create_sheet("8 - Play by Play")
    ws8.sheet_view.showGridLines = True
    ws8.auto_filter.ref = f"A1:{get_column_letter(len(p_data.columns))}{len(p_data)+1}"

    header_row(ws8, 1, list(p_data.columns), bg=DARK)
    for i, row_data in enumerate(p_data.values, 2):
        bg = GRAY if i % 2 == 0 else WHITE
        for j, val in enumerate(row_data, 1):
            c = ws8.cell(row=i, column=j, value=val)
            c.fill      = make_fill(bg)
            c.font      = Font(color='000000', size=9, name='Calibri')
            c.alignment = make_align()
            c.border    = make_border()

    for col_idx in range(1, len(p_data.columns) + 1):
        set_col_width(ws8, col_idx, 14)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()



# ============================================================
# HELPER FUNCTIONS
# ============================================================

def classify_leverage(dn, dist, yard_ln=None):
    try:
        d, y = int(dn), int(dist)
    except (ValueError, TypeError):
        return "Unknown", 0
    if d == 1:
        if y <= 5:    base = 2
        elif y <= 10: base = 1
        else:         base = -1
    elif d == 2:
        if y <= 3:    base = 2
        elif y <= 7:  base = 1
        else:         base = -1
    elif d in (3, 4):
        if y <= 2:    base = 2
        elif y <= 6:  base = 0
        else:         base = -2
    else:
        return "Unknown", 0
    modifier = 0.0
    if yard_ln is not None:
        try:
            yl = int(yard_ln)
            if 1 <= yl <= 20:  modifier = +0.5
            elif yl <= -30:    modifier = -0.5
        except (ValueError, TypeError):
            pass
    score = base + modifier
    if score >= 1.5:    band = "High"
    elif score >= 0.5:  band = "Med"
    elif score >= -0.5: band = "Neutral"
    else:               band = "Low"
    return band, round(score, 1)


def process_offensive_logic(formation):
    f = str(formation).upper().strip()
    match = re.match(r'^(\d)(\d)', f)
    if match:
        return f"{match.group(1)}{match.group(2)}"
    if any(x in f for x in ["HEAVY", "JUMBO", "BIG"]): return "23"
    if "EMPTY" in f:                                     return "00"
    if "DOUBLE Y DOUBLE WING" in f:                      return "13"
    if "TREY" in f:                                      return "12"
    if "DUBS" in f or "TRIPS" in f:                      return "10"
    if "SPREAD" in f or "WING" in f:                     return "11"
    return "11"


def get_stars(pct):
    if pct >= 85: return ""
    if pct >= 75: return ""
    if pct >= 65: return ""
    if pct >= 50: return ""
    return ""


def dls_grade(x):
    if x >= 1.5:  return "A"
    if x >= 0.8:  return "B"
    if x >= 0.2:  return "C"
    if x >= -0.5: return "D"
    return "F"


EXPECTED_GAIN = {
    (1,'1-5'):3.5,(1,'6-10'):4.2,(1,'11+'):3.0,
    (2,'1-3'):3.0,(2,'4-7'):4.5,(2,'8+'):5.5,
    (3,'1-2'):2.5,(3,'3-6'):5.0,(3,'7+'):7.0,
    (4,'1-2'):2.0,(4,'3+'):5.0,
}

def dist_bucket(dn, dist):
    d, y = int(dn), int(dist)
    if d == 1:
        if y <= 5: return (1,'1-5')
        elif y <= 10: return (1,'6-10')
        else: return (1,'11+')
    elif d == 2:
        if y <= 3: return (2,'1-3')
        elif y <= 7: return (2,'4-7')
        else: return (2,'8+')
    elif d == 3:
        if y <= 2: return (3,'1-2')
        elif y <= 6: return (3,'3-6')
        else: return (3,'7+')
    elif d == 4:
        if y <= 2: return (4,'1-2')
        else: return (4,'3+')
    return None


def build_sss(p_data, cols):
    df_s = p_data.copy().reset_index(drop=True)
    stress = df_s[(df_s[cols['dn']] == 3) & (df_s[cols['dist']] >= 5)]
    causes = []
    for idx in stress.index:
        if idx > 0:
            prev = df_s.loc[idx - 1]
            causes.append({
                'Stress_Situation': f"Third & {df_s.loc[idx, cols['dist']]}",
                'Caused_By_Play':   prev[cols['play']],
                'Caused_By_Type':   prev[cols['type']],
                'Caused_By_Form':   prev[cols['form']],
                'Prior_Gain':       prev[cols['gain']],
                'Prior_Result':     prev[cols['result']],
            })
    sss_df = pd.DataFrame(causes)
    if not sss_df.empty:
        sss_summary = (
            sss_df.groupby('Caused_By_Type')
            .agg(Stress_Plays=('Caused_By_Type','count'), Avg_Prior_Gain=('Prior_Gain','mean'))
            .round(1)
        )
        sss_summary['Stress %'] = (
            sss_summary['Stress_Plays'] / sss_summary['Stress_Plays'].sum() * 100
        ).round(0).astype(int)
        sss_by_form = (
            sss_df.groupby('Caused_By_Form')
            .agg(Stress_Count=('Caused_By_Form','count'))
            .sort_values('Stress_Count', ascending=False).head(8)
        )
    else:
        sss_summary = pd.DataFrame()
        sss_by_form = pd.DataFrame()
    return sss_df, sss_summary, sss_by_form

def build_fpar(p_data, cols):
    df_p = p_data.copy()
    def field_zone(yl):
        y = int(yl)
        if y <= -30:  return "Backed Up (own 30-)"
        elif y <= 0:  return "Own Territory (30-50)"
        elif y <= 20: return "Opp Territory (50-opp30)"
        else:         return "Scoring Zone (opp 20+)"
    df_p['Field_Zone'] = df_p[cols['field']].apply(field_zone)
    df_p['Is_Pass']    = (df_p[cols['type']] == 'PASS').astype(int)
    fpar_df = (
        df_p.groupby(['Field_Zone', cols['dn']])
        .agg(Plays=('Is_Pass','count'), Pass_Rate=('Is_Pass','mean'),
             Avg_Gain=(cols['gain'],'mean'), Success_Rate=('Is_Succ','mean'), FD_Rate=('Is_FD','mean'))
        .round(3)
    )
    fpar_df['Pass_Rate']    = (fpar_df['Pass_Rate']    * 100).round(0).astype(int)
    fpar_df['Success_Rate'] = (fpar_df['Success_Rate'] * 100).round(0).astype(int)
    fpar_df['FD_Rate']      = (fpar_df['FD_Rate']      * 100).round(0).astype(int)
    fpar_df['Avg_Gain']     = fpar_df['Avg_Gain'].round(1)
    zone_order = {
        "Backed Up (own 30-)":1,"Own Territory (30-50)":2,
        "Opp Territory (50-opp30)":3,"Scoring Zone (opp 20+)":4
    }
    fpar_df['Zone_Order'] = fpar_df.index.get_level_values('Field_Zone').map(zone_order)
    return fpar_df.sort_values(['Zone_Order', cols['dn']]).drop(columns='Zone_Order')


def build_intel(p_data, df, cols):
    intel = []
    total = len(p_data)

    # 1. Post-sack tendency
    if cols['result'] in df.columns:
        sack_mask = (
            df[cols['result']].str.contains('Sack', case=False, na=False) |
            ((df[cols['type']] == 'PASS') & (df[cols['gain']] <= -4))
        )
        post_sack = df.loc[[i+1 for i in df[sack_mask].index if i+1 in df.index]]
        if not post_sack.empty:
            rate = round((post_sack[cols['type']].str.upper() == 'RUN').mean() * 100)
            intel.append({
                "Category": "Sequence",
                "Signal": "Post-Sack Play Call",
                "Stat": f"{rate}pct RUN",
                "Coaching Note": "They run to get back on schedule after a sack  apply run blitz immediately after pressure" if rate >= 60 else "They keep passing after a sack  bring pressure again, they wont adjust",
            })

    # 2. First down pass rate
    first_downs = p_data[p_data[cols['dn']] == 1]
    if not first_downs.empty:
        fd_pass_rate = round((first_downs[cols['type']] == 'PASS').mean() * 100)
        intel.append({
            "Category": "Tendency",
            "Signal": "First Down Pass Rate",
            "Stat": f"{fd_pass_rate}pct",
            "Coaching Note": "Pass-first on early downs  show two-high shell to bait short completions, then rally" if fd_pass_rate >= 55 else "Run-first on First down  load the box, force them to prove they can pass",
        })

    # 3. Third & short conversion
    third_short = p_data[(p_data[cols['dn']] == 3) & (p_data[cols['dist']] <= 3)]
    if len(third_short) >= 3:
        conv_rate = round(third_short['Is_FD'].mean() * 100)
        intel.append({
            "Category": "Efficiency",
            "Signal": "Third & Short Conversion (13 yds)",
            "Stat": f"{conv_rate}pct",
            "Coaching Note": "Dangerous in short-yardage  commit extra defender at line of scrimmage" if conv_rate >= 70 else "Stoppable in short-yardage  they struggle to get the tough yards when it matters",
        })

    # 4. Motion rate and delta
    if cols['motion'] in p_data.columns:
        motion_plays = p_data[
            p_data[cols['motion']].notna() &
            (p_data[cols['motion']].astype(str).str.strip() != '')
        ]
        motion_rate = round(len(motion_plays) / total * 100) if total else 0
        if not motion_plays.empty:
            motion_succ    = round(motion_plays['Is_Succ'].mean() * 100)
            no_motion_succ = round(p_data[~p_data.index.isin(motion_plays.index)]['Is_Succ'].mean() * 100)
            delta = motion_succ - no_motion_succ
            intel.append({
                "Category": "Scheme",
                "Signal": "Pre-Snap Motion Rate",
                "Stat": f"{motion_rate}pct of plays",
                "Coaching Note": f"Motion adds +{delta}pct success  disrupt at the snap, dont let them get free releases" if delta >= 5 else f"Motion not meaningfully helping ({delta}pct)  dont overreact to motion keys",
            })

    # 5. Explosive dependency
    exp_rate    = round(p_data['Is_Explosive'].mean() * 100)
    non_exp_avg = round(p_data[p_data['Is_Explosive'] == 0][cols['gain']].mean(), 1)
    intel.append({
        "Category": "Identity",
        "Signal": "Explosive Play Dependency",
        "Stat": f"{exp_rate}pct of plays 15 yds",
        "Coaching Note": f"Big-play dependent  eliminate explosives and their non-explosive avg drops to {non_exp_avg} yds/play" if exp_rate >= 12 else f"Not big-play dependent  they grind consistently ({non_exp_avg} yds/play without explosives)",
    })

    # 6. Red/green zone pass rate
    rz = p_data[p_data[cols['field']].between(1, 20)]
    if len(rz) >= 4:
        rz_pass_rate = round((rz[cols['type']] == 'PASS').mean() * 100)
        rz_succ      = round(rz['Is_Succ'].mean() * 100)
        intel.append({
            "Category": "Red Zone",
            "Signal": "Scoring Zone Pass Rate (inside 20)",
            "Stat": f"{rz_pass_rate}pct PASS | {rz_succ}pct success",
            "Coaching Note": "Pass-heavy in scoring position  play press man, disrupt route timing" if rz_pass_rate >= 55 else "Run-heavy in scoring position  stack the box, force them to throw it in",
        })

    # 7. First down plays creating Second & long
    if not first_downs.empty:
        created_Second_long = round((first_downs[cols['gain']] <= 2).mean() * 100)
        intel.append({
            "Category": "Self-Scout",
            "Signal": "First Downs Ending in 2 Yd Gain",
            "Stat": f"{created_Second_long}pct",
            "Coaching Note": "They frequently strand themselves  win First down and the drive often stalls on its own" if created_Second_long >= 35 else "Efficient on First down  dont give them easy early-down gains",
        })

    # 8. Interception rate
    pass_plays = len(p_data[p_data[cols['type']] == 'PASS'])
    if pass_plays >= 5:
        int_per_pass = round(p_data['Is_Int'].sum() / pass_plays * 100, 1)
        intel.append({
            "Category": "Turnover",
            "Signal": "Interception Rate (per pass attempt)",
            "Stat": f"{int_per_pass}pct",
            "Coaching Note": "Turnover-prone passer  force obvious passing situations and play the sticks" if int_per_pass >= 5 else "Ball-secure passer  dont gamble on picks, play assignment defense",
        })

    return pd.DataFrame(intel) if intel else pd.DataFrame()


# ============================================================
# SCOUT REPORT GENERATOR
# ============================================================

def generate_scout_report(p_data, drive_dla, pers_dla,
                           fpar_df, sss_summary, sss_by_form,
                           chain, cols):
    lines = []
    total    = len(p_data)
    runs     = (p_data[cols['type']] == 'RUN').sum()
    passes   = (p_data[cols['type']] == 'PASS').sum()
    run_pct  = round(runs / total * 100) if total else 0
    pass_pct = round(passes / total * 100) if total else 0
    avg_gain = round(p_data[cols['gain']].mean(), 1)
    fd_rate  = round(p_data['Is_FD'].mean() * 100)
    succ_rt  = round(p_data['Is_Succ'].mean() * 100)
    exp_rt   = round(p_data['Is_Explosive'].mean() * 100)
    avg_dls  = round(drive_dla['DLS'].mean(), 2) if not drive_dla.empty else 0
    dls_g    = dls_grade(avg_dls)

    if run_pct >= 60:
        identity = f"a run-heavy offense ({run_pct}pct run rate)"
    elif pass_pct >= 60:
        identity = f"a pass-heavy offense ({pass_pct}pct pass rate)"
    else:
        identity = f"a balanced offense ({run_pct}pct run / {pass_pct}pct pass)"

    _ident_tail = "a dangerous big-play threat." if exp_rt >= 15 else "not a big-play threat - bend-dont-break works."
    _ident_chain = "consistently stay ahead of the chains" if succ_rt >= 55 else "frequently fall behind the chains"
    lines.append((" Offensive Identity",
        f"This opponent runs {identity} averaging {avg_gain} yards per play. "
        f"First Down Rate: {fd_rate}pct, Success Rate: {succ_rt}pct. "
        f"They {_ident_chain}. "
        f"Explosive plays (15+ yards): {exp_rt}pct of offense. {_ident_tail}"
    ))

    if not drive_dla.empty:
        best_drive  = drive_dla['DLS'].max()
        worst_drive = drive_dla['DLS'].min()
        pct_a_b = round((drive_dla['DLS_Grade'].isin(['A','B'])).mean() * 100)
        pct_d_f = round((drive_dla['DLS_Grade'].isin(['D','F'])).mean() * 100)
        _dls_note = "Exploit: Force early negative plays - their low-leverage drives collapse quickly." if avg_dls < 0.8 else "Caution: This offense controls drives well - requires consistent TFLs on first down."
        lines.append((" Drive Control (DLS)",
            f"Avg Drive Leverage Score: {avg_dls} (Grade: {dls_g}). "
            f"{pct_a_b}pct of drives graded A or B. "
            f"{pct_d_f}pct of drives graded D or F. "
            f"Best DLS: {best_drive} | Worst: {worst_drive}. "
            f"{_dls_note}"
        ))

    if not pers_dla.empty:
        top_pers       = pers_dla.sort_values('Plays', ascending=False)
        primary        = top_pers.index[0] if len(top_pers) > 0 else "N/A"
        primary_pct    = round(int(top_pers.iloc[0]['Plays']) / total * 100) if total else 0
        primary_dls    = top_pers.iloc[0]['DLS']
        primary_run    = top_pers.iloc[0]['Run%']
        primary_pass   = top_pers.iloc[0]['Pass%']
        qual           = pers_dla[pers_dla['Plays'] >= 5]
        worst_pers_row = qual.sort_values('DLS').iloc[0] if len(qual) > 0 else None
        best_pers_row  = qual.sort_values('DLS', ascending=False).iloc[0] if len(qual) > 0 else None
        worst_note = f"Their {worst_pers_row.name} personnel has the lowest DLS ({worst_pers_row['DLS']})  when they align here, stress situations follow." if worst_pers_row is not None else ""
        best_note  = f"Their {best_pers_row.name} personnel is their most controlled grouping (DLS: {best_pers_row['DLS']})  expect this on critical downs." if best_pers_row is not None else ""
        lines.append((" Personnel Tendencies",
            f"Primary group: {primary} ({primary_pct}pct of plays, {primary_run}pct run / {primary_pass}pct pass, DLS: {primary_dls}). "
            f"{best_note} "
            f"{worst_note} "
            "Exploit: When their low-DLS personnel aligns, apply pressure - dont give up the conversion."
        ))

    if not sss_summary.empty:
        top_cause  = sss_summary.sort_values('Stress %', ascending=False).iloc[0]
        cause_type = top_cause.name
        cause_pct  = int(top_cause['Stress %'])
        cause_gain = round(top_cause['Avg_Prior_Gain'], 1)
        form_note  = ""
        if not sss_by_form.empty:
            tsf = sss_by_form.index[0]
            tsc = int(sss_by_form.iloc[0]['Stress_Count'])
            form_note = f"Formation most responsible: {tsf} ({tsc} stress situations generated)."
        lines.append(("Stress Pattern Analysis (SSS)",
            f"{cause_pct}pct of their Third-and-long situations created by {cause_type} plays. "
            f"Averaging only {cause_gain} yards on the prior snap. "
            f"{form_note} "
            f"Exploit: Stop their {cause_type.lower()} game on early downs. "
            f"Hold them below {cause_gain + 1} yards on First and Second down."
        ))

    if not fpar_df.empty:
        fpar_reset  = fpar_df.reset_index()
        bu_First      = fpar_reset[(fpar_reset['Field_Zone'] == 'Backed Up (own 30-)') & (fpar_reset[cols['dn']] == 1)]
        sc_First      = fpar_reset[(fpar_reset['Field_Zone'] == 'Scoring Zone (opp 20+)') & (fpar_reset[cols['dn']] == 1)]
        backed_note = ""
        if not bu_First.empty:
            bu_pass = int(bu_First.iloc[0]['Pass_Rate'])
            bu_succ = int(bu_First.iloc[0]['Success_Rate'])
            backed_note = f"When backed up on their own 30 or deeper, they pass {bu_pass}pct on First down ({bu_succ}pct success)  {'predictable and stoppable' if bu_pass >= 55 else 'they run it safe  limit your risk in this zone too'}."
        scoring_note = ""
        if not sc_First.empty:
            sc_pass = int(sc_First.iloc[0]['Pass_Rate'])
            sc_succ = int(sc_First.iloc[0]['Success_Rate'])
            scoring_note = f"Inside your scoring zone, they pass {sc_pass}pct on First down ({sc_succ}pct success)  {'get physical at the line, disrupt route timing' if sc_pass >= 55 else 'stack the box, they want to run it in from here'}."
        lines.append((" Field Position Tendencies",
            f"{backed_note} "
            f"{scoring_note} "
            "Exploit: Zone-by-zone tendencies let you call the right front before the snap."
        ))

    if not chain.empty:
        top_plays_text = "\n".join([f"- {play}  {row['Plays']} plays, {row['FD Rate %']}pct FD, {row['Success Rate %']}pct success" for play, row in chain.head(5).iterrows()])
        bot_plays_text = "\n".join([f"- {play}  {row['Plays']} plays, {row['FD Rate %']}pct FD" for play, row in chain.sort_values('FD Rate %').head(3).iterrows()])
        lines.append((" Chain-Moving Plays to Stop",
            f"Most dangerous chain-movers:\n{top_plays_text}\n\n"
            f"Frequent calls with low conversion:\n{bot_plays_text}\n\n"
            "Exploit: Take away their top chain-movers and they fall back on low-conversion habits."
        ))

    verdict_score = 0

    if fd_rate >= 30:   verdict_score += 1
    if fd_rate >= 38:   verdict_score += 1
    if succ_rt >= 45:   verdict_score += 1
    if succ_rt >= 52:   verdict_score += 1
    if exp_rt >= 10:    verdict_score += 1
    if exp_rt >= 15:    verdict_score += 1
    if avg_dls >= 0.3:  verdict_score += 1
    if avg_dls >= 0.7:  verdict_score += 1
    if avg_gain >= 5.0: verdict_score += 1
    if avg_gain >= 6.5: verdict_score += 1

    # Identify dominant signals
    strengths = []
    if avg_gain >= 6.5:   strengths.append(f"elite yards per play ({avg_gain})")
    elif avg_gain >= 5.0: strengths.append(f"strong yards per play ({avg_gain})")
    if exp_rt >= 15:      strengths.append(f"dangerous big-play rate ({exp_rt}pct)")
    elif exp_rt >= 10:    strengths.append(f"moderate explosive threat ({exp_rt}pct)")
    if succ_rt >= 52:     strengths.append(f"elite success rate ({succ_rt}pct)")
    elif succ_rt >= 45:   strengths.append(f"above-average success rate ({succ_rt}pct)")
    if fd_rate >= 50:     strengths.append(f"high first down rate ({fd_rate}pct)")
    elif fd_rate >= 38:   strengths.append(f"solid first down rate ({fd_rate}pct)")
    if not drive_dla.empty:
        dls = round(drive_dla['DLS'].mean(), 2)
        if dls >= 0.7:    strengths.append(f"controlled drive leverage (DLS: {dls})")
        elif dls >= 0.3:  strengths.append(f"moderate drive control (DLS: {dls})")

    strength_text = ", ".join(strengths[:3]) if strengths else "balanced production"

    if verdict_score >= 7:
        verdict = f" High-Threat Offense. Powered by {strength_text}. No single silver bullet  must stop them consistently on every snap."
    elif verdict_score >= 4:
        verdict = f" Moderate-Threat Offense. Key weapons: {strength_text}. Attack their stress patterns on early downs and force them behind the chains."
    else:
        verdict = f" Manageable Offense. Limited by {strength_text if strengths else 'inconsistent execution'}. Force early-down stops and let their tendencies beat them."
    lines.append((" Overall Scouting Verdict", verdict))
    return lines


# ============================================================
# PAGE CONFIG
# ============================================================

st.set_page_config(page_title="FormationIQ", page_icon="", layout="wide")

with st.sidebar:
    logo_files = ["Logo.png", "logo.png"]
    found_logo = False
    for lf in logo_files:
        if os.path.exists(lf):
            st.image(lf, width=150)
            found_logo = True
            break
    if not found_logo:
        st.subheader(" CARLSBAD FOOTBALL")
    st.write("-")
    st.caption("FormationIQ v9.0  Final Build")

st.title(" FormationIQ  Offensive Scouting Analytics")
st.markdown("""
> Upload a Hudl play-by-play export and instantly break down your opponents offensive tendencies  by formation, personnel, down, distance, and more.
""")

col1, col2, col3 = st.columns(3)

with col1:
    st.markdown("###  Getting Started")
    st.markdown("""
- Export your opponents playlist from Hudl as CSV or Excel
- Upload the file using the uploader below
- All tabs populate automatically  no setup needed
""")

with col2:
    st.markdown("###  Whats Inside")
    st.markdown("""
- Formation & Personnel breakdowns
- Run/Pass tendencies by down & distance
- Field zone and hash analysis
- Play success and explosive play rates
- Custom Pivot Lab  build your own views
""")

with col3:
    st.markdown("###  Tips")
    st.markdown("""
- Use the Play Type filter to isolate run or pass
- All charts and tables are exportable to Excel
- Try the sample data below if you dont have a file yet
""")

st.divider()
uploaded_file = st.file_uploader("Upload Hudl file (CSV or Excel)", type=["csv", "xlsx"])

with st.expander(" No file? Download sample data"):
    sample_csv = """PLAY #,ODK,DN,DIST,YARD LN,HASH,OFF FORM,OFF STR,OFF PLAY,PLAY TYPE,GN/LS,RESULT,BACKFIELD,EFF,TARGET,MOTION DIR,PLAY DIR
1,O,1,10,35,M,20 Wing,R,QUICK PASS,PASS,7,Complete,,Y,,,R
2,O,2,3,28,R,20 Wing,L,ZONE,RUN,4,Rush,,Y,,,L
3,O,1,10,24,L,DUBS,BAL,DROP BACK PASS,PASS,0,Incomplete,,N,,,R
4,O,2,10,24,L,DUBS,BAL,QB BLAST,RUN,12,Rush,,Y,,,L
5,O,1,10,12,R,11 spread,R,PLAY ACTION PASS,PASS,12,Complete,,Y,,,R
6,O,1,10,38,M,20 Wing,L,WIDE ZONE,RUN,34,Rush,,Y,,,R
7,O,3,5,20,R,DUBS,BAL,QUICK PASS,PASS,0,Incomplete,,N,,,L
8,O,3,5,20,R,20 Wing,L,COUNTER H,RUN,8,Rush,,Y,,,R
9,O,1,10,8,L,20 Wing,R,DIVE,RUN,8,Rush TD,,Y,,,L
10,O,1,10,-22,R,10 trips,R,BUBBLE,PASS,50,Complete TD,,Y,,,R
11,O,2,7,-36,R,11 spread wing,R,LONG TRAP,RUN,1,Rush,,N,,,L
12,O,3,6,-37,M,20 Wing,L,DROP BACK PASS,PASS,0,Incomplete,,N,,,R"""
    st.download_button(
        label=" Download Sample CSV",
        data=sample_csv,
        file_name="sample_hudl_data.csv",
        mime="text/csv"
    )

# ============================================================
# MAIN APP
# ============================================================

if uploaded_file:
    if uploaded_file.name.lower().endswith(".csv"):
        df = pd.read_csv(uploaded_file)
    else:
        df = pd.read_excel(uploaded_file)
    cols = {
        'type':   'PLAY TYPE',
        'form':   'OFF FORM',
        'gain':   'GN/LS',
        'dn':     'DN',
        'dist':   'DIST',
        'play':   'OFF PLAY',
        'field':  'YARD LN',
        'odk':    'ODK',
        'hash':   'HASH',
        'p_dir':  'PLAY DIR',
        'motion': 'MOTION DIR',
        'result': 'RESULT',
    }
    
    if all(cols[k] in df.columns for k in ['type', 'form', 'gain']):

        df[cols['type']]  = df[cols['type']].astype(str).str.upper().str.strip()
        df[cols['gain']]  = pd.to_numeric(df[cols['gain']],  errors='coerce').fillna(0).round(0).astype(int)
        df[cols['dn']]    = pd.to_numeric(df[cols['dn']],    errors='coerce').fillna(0).astype(int)
        df[cols['dist']]  = pd.to_numeric(df[cols['dist']],  errors='coerce').fillna(0).astype(int)
        df[cols['field']] = pd.to_numeric(df[cols['field']], errors='coerce').fillna(0).astype(int)
        df['Drive_ID']    = (df[cols['odk']] != df[cols['odk']].shift()).cumsum()

        p_data = df[df[cols['type']].isin(['RUN', 'PASS'])].copy()
        p_data['PERSONNEL'] = p_data[cols['form']].apply(process_offensive_logic)
        p_data['Is_FD']     = (p_data[cols['gain']] >= p_data[cols['dist']]).astype(int)
        p_data['Is_Int']    = p_data[cols['result']].str.contains('Interception', case=False, na=False).astype(int)

        def calc_succ(row):
            d, dist, g = row[cols['dn']], row[cols['dist']], row[cols['gain']]
            if d == 1: return g >= (dist * 0.45)
            if d == 2: return g >= (dist * 0.65)
            return g >= dist

        p_data['Is_Succ']      = p_data.apply(calc_succ, axis=1).astype(int)
        p_data['Is_Explosive'] = (p_data[cols['gain']] >= 15).astype(int)

        leva = p_data.apply(
            lambda r: classify_leverage(r[cols['dn']], r[cols['dist']], r[cols['field']]), axis=1
        )
        p_data['Leverage_Band']  = leva.apply(lambda x: x[0])
        p_data['Leverage_Score'] = leva.apply(lambda x: x[1])

        if 'Drive_ID' in df.columns and 'Drive_ID' not in p_data.columns:
            p_data = p_data.merge(df[['PLAY #', 'Drive_ID']], on='PLAY #', how='left')

        drive_view = p_data.dropna(subset=['Drive_ID']).copy()

        drive_dla = drive_view.groupby('Drive_ID').agg(
            Plays=('Leverage_Score','count'), DLS=('Leverage_Score','mean'),
            FD_Rate=('Is_FD','mean'), Success_Rate=('Is_Succ','mean'), Explosive_Rt=('Is_Explosive','mean'),
        ).round(2)
        drive_dla['High_Lev%']    = drive_view.groupby('Drive_ID')['Leverage_Score'].apply(lambda x: round((x >= 1.5).mean() * 100))
        drive_dla['Low_Lev%']     = drive_view.groupby('Drive_ID')['Leverage_Score'].apply(lambda x: round((x <= -1).mean() * 100))
        drive_dla['FD_Rate']      = (drive_dla['FD_Rate'] * 100).round(0).astype(int)
        drive_dla['Success_Rate'] = (drive_dla['Success_Rate'] * 100).round(0).astype(int)
        drive_dla['Explosive_Rt'] = (drive_dla['Explosive_Rt'] * 100).round(0).astype(int)
        drive_dla['DLS_Grade']    = drive_dla['DLS'].apply(dls_grade)

        pers_dla = p_data.groupby('PERSONNEL').agg(
            Plays=('Leverage_Score','count'), DLS=('Leverage_Score','mean'),
            Avg_Gain=(cols['gain'],'mean'), FD_Rate=('Is_FD','mean'),
            Success_Rate=('Is_Succ','mean'), Explosive_Rt=('Is_Explosive','mean'),
        ).round(2)
        pers_dla['High_Lev%']    = p_data.groupby('PERSONNEL')['Leverage_Score'].apply(lambda x: round((x >= 1.5).mean() * 100))
        pers_dla['Low_Lev%']     = p_data.groupby('PERSONNEL')['Leverage_Score'].apply(lambda x: round((x <= -1).mean() * 100))
        pers_dla['Run%']         = p_data.groupby('PERSONNEL')[cols['type']].apply(lambda x: round((x == 'RUN').mean() * 100))
        pers_dla['Pass%']        = p_data.groupby('PERSONNEL')[cols['type']].apply(lambda x: round((x == 'PASS').mean() * 100))
        pers_dla['FD_Rate']      = (pers_dla['FD_Rate'] * 100).round(0).astype(int)
        pers_dla['Success_Rate'] = (pers_dla['Success_Rate'] * 100).round(0).astype(int)
        pers_dla['Explosive_Rt'] = (pers_dla['Explosive_Rt'] * 100).round(0).astype(int)
        pers_dla['Avg_Gain']     = pers_dla['Avg_Gain'].round(1)
        pers_dla['DLS_Grade']    = pers_dla['DLS'].apply(dls_grade)

        pf_dla = p_data.groupby(['PERSONNEL', cols['form']]).agg(
            Plays=('Leverage_Score','count'), DLS=('Leverage_Score','mean'),
            Avg_Gain=(cols['gain'],'mean'), FD_Rate=('Is_FD','mean'),
            Success_Rate=('Is_Succ','mean'), Explosive_Rt=('Is_Explosive','mean'),
        ).round(2)
        pf_dla['High_Lev%']    = p_data.groupby(['PERSONNEL', cols['form']])['Leverage_Score'].apply(lambda x: round((x >= 1.5).mean() * 100))
        pf_dla['Low_Lev%']     = p_data.groupby(['PERSONNEL', cols['form']])['Leverage_Score'].apply(lambda x: round((x <= -1).mean() * 100))
        pf_dla['FD_Rate']      = (pf_dla['FD_Rate'] * 100).round(0).astype(int)
        pf_dla['Success_Rate'] = (pf_dla['Success_Rate'] * 100).round(0).astype(int)
        pf_dla['Explosive_Rt'] = (pf_dla['Explosive_Rt'] * 100).round(0).astype(int)
        pf_dla['Avg_Gain']     = pf_dla['Avg_Gain'].round(1)
        pf_dla['DLS_Grade']    = pf_dla['DLS'].apply(dls_grade)
        pf_dla = pf_dla[pf_dla['Plays'] >= 5]
        sss_df, sss_summary, sss_by_form = build_sss(p_data, cols)
        if 'CausedByForm' in sss_df.columns:
            sss_df['CausedByForm'] = sss_df['CausedByForm'].astype(str)
        if cols['form'] in sss_df.columns:
            sss_df[cols['form']] = sss_df[cols['form']].astype(str)
        if not sss_by_form.empty:
            sss_by_form.index = sss_by_form.index.astype(str)
        fpar_df = build_fpar(p_data, cols)
        intel_df = build_intel(p_data, df, cols)
        chain = p_data.groupby(cols['play'])['Is_FD'].agg(['sum','count'])
        chain.columns = ['First Downs','Plays']
        chain['FD Rate %']      = (chain['First Downs'] / chain['Plays'] * 100).round(0).astype(int)
        chain['Success Rate %'] = p_data.groupby(cols['play'])['Is_Succ'].mean().mul(100).round(0).astype(int)
        chain = chain[chain['Plays'] >= 3].sort_values('FD Rate %', ascending=False).head(15)


        pers_counts = p_data['PERSONNEL'].value_counts().to_frame("Plays")
        pers_counts['%'] = (pers_counts['Plays'] / pers_counts['Plays'].sum() * 100).round(0).astype(int)

        t3 = p_data[p_data[cols['dn']] == 3].copy()
        if not t3.empty:
            t3['Sit'] = t3[cols['dist']].apply(
                lambda x: "Third & Short (1-3)" if x <= 3 else ("Third & Mid (4-7)" if x <= 7 else "Third & Long (7+)")
            )
            t3_summary = t3.groupby('Sit')['Is_FD'].mean().mul(100).round(0).astype(int).to_frame("FD Rate %")
        else:
            t3_summary = pd.DataFrame()

        scout_sections = generate_scout_report(
            p_data, drive_dla, pers_dla,
            fpar_df, sss_summary, sss_by_form, chain, cols
        )
        verdict_score = 0
        if round(p_data['Is_FD'].mean()*100) >= 30:        verdict_score += 1
        if round(p_data['Is_FD'].mean()*100) >= 38:        verdict_score += 1
        if round(p_data['Is_Succ'].mean()*100) >= 45:      verdict_score += 1
        if round(p_data['Is_Succ'].mean()*100) >= 52:      verdict_score += 1
        if round(p_data['Is_Explosive'].mean()*100) >= 10: verdict_score += 1
        if round(p_data['Is_Explosive'].mean()*100) >= 15: verdict_score += 1
        if (not drive_dla.empty) and drive_dla['DLS'].mean() >= 0.3: verdict_score += 1
        if (not drive_dla.empty) and drive_dla['DLS'].mean() >= 0.7: verdict_score += 1
        if round(p_data[cols['gain']].mean(), 1) >= 5.0:   verdict_score += 1
        if round(p_data[cols['gain']].mean(), 1) >= 6.5:   verdict_score += 1
        export_options = {
            "Personnel Identity":         pers_counts,
            "Third Down Summary":           t3_summary,
            "Chain Moving":               chain,
            "Drive Leverage-Per Drive":   drive_dla,
            "Drive Leverage-Personnel":   pers_dla,
            "Drive Leverage-Pers+Form":   pf_dla,
            "Sequence Stress Score":      sss_summary,
            "Stress by Formation":        sss_by_form,
            "Field Position Aggression":  fpar_df.reset_index(),
            "AI Scouting Intelligence":   intel_df,
        }
        
        #  SIDEBAR 
        with st.sidebar:
            st.markdown("###  Game Summary")
            st.metric("Total Plays",  len(p_data))
            st.metric("Run Plays",    len(p_data[p_data[cols['type']] == 'RUN']))
            st.metric("Pass Plays",   len(p_data[p_data[cols['type']] == 'PASS']))
            st.metric("Avg Gain",     f"{p_data[cols['gain']].mean():.1f} yds")
            st.metric("FD Rate",      f"{round(p_data['Is_FD'].mean()*100)}pct")
            st.metric("Success Rate", f"{round(p_data['Is_Succ'].mean()*100)}pct")
            st.write("-")
            st.subheader(" Download Full Report")
            excel_data = build_excel_export(
                export_options, p_data, drive_dla, pers_dla,
                fpar_df, sss_summary,
                sss_by_form, chain, intel_df, scout_sections,
                cols, verdict_score
            )
            st.download_button(
                label=" Download FormationIQ Report",
                data=excel_data,
                file_name="FormationIQ_Report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )


        #  TABS 
        
        tabs = st.tabs([
            "Definitions", "Personnel Identity", "Third Down Efficiency",
            "Chain Moving", "Red/Green Zone", "Winning Probability",
            "Pivot Lab", "Vic Fangio", "Drive Leverage (DLA)"
        ])
        #  TAB 0: DEFINITIONS 
        with tabs[0]:
            st.header(" Metric Definitions")
            st.caption("Reference guide for every metric used in FormationIQ.")

            st.subheader(" First Down Rate (FD Rate)")
            st.markdown("""
How often this offense picks up the first down marker  gain equals or exceeds the distance needed.

Examples:
Second & 7  gain of 8 yards  First down counted
Second & 7  gain of 6 yards  Not counted
Third & 4  gain of 4 yards  First down counted

> Below 35% = chains rarely moving. Above 50% = very difficult to get off the field.
""")
            st.divider()

            st.subheader(" Success Rate")
            st.markdown("""
A more precise measure than FD Rate. A play is "successful" if it gained
enough for the situation  not just whether it converted.

| Down | Threshold | Example |
|-|-|-|
| First |  45% of distance | First & 10  need 4.5 yds  5 yd gain  |
| First |  45% of distance | First & 10  need 4.5 yds  3 yd gain  |
| Second |  65% of distance | Second & 8  need 5.2 yds  6 yd gain  |
| Second |  65% of distance | Second & 8  need 5.2 yds  4 yd gain  |
| Third/Fourth | Full conversion | Third & 6  need 6 yds  5 yd gain  |

Why it matters more than FD Rate:
A team can have a 42% FD Rate but 38% Success Rate  they converted
first downs but were constantly behind the chains before doing so.
Success Rate exposes that theyre surviving on Third down luck rather than
building drives consistently.

> Above 50% = disciplined chain-moving offense. Below 40% = living and dying by big plays.
""")
            st.divider()

            st.subheader(" Drive Leverage Score (DLS)")
            st.markdown("""
Measures how much control the offense had at every snap.

| Situation | Score |
|-|-|
| First & 5, Second & 3, Third/Fourth & 1-2 | +2 (High) |
| Normal First & 10, Second & 4-7 | +1 (Med) |
| First/Second behind chains | -1 (Low) |
| Third/Fourth & 7+ | -2 (Stress) |

Field position modifier: Red zone +0.5 | Backed up own 30 -0.5

Grade: A  1.5 | B  0.8 | C  0.2 | D  -0.5 | F < -0.5

Example: A drive with plays on First & 10 (+1), Second & 3 (+2), Third & 1 (+2) = avg DLS of 1.67  Grade A.
A drive with First & 10 (+1), incomplete pass  Second & 10 (-1), Third & 10 (-2) = avg DLS of -0.67  Grade F.
""")
            st.divider()

            st.subheader(" Sequence Stress Score (SSS)")
            st.markdown("""
Tracks how often the offense enters Third & 5+ situations and identifies
which prior play type or formation caused the stress.

Example:
First & 10  incomplete pass (0 yds)  Second & 10  run for 2 yds  Third & 8  stress situation
- SSS tags the Second down run as the cause of the Third & 8
> Use this to find the root cause of drive breakdowns  not just the symptom.
""")
            st.divider()

            st.subheader(" Field Position Aggression Rating (FPAR)")
            st.markdown("""
Pass rate, success rate, and avg gain by field zone and down.

| Zone | Description |
|-|-|
| Backed Up | Own 30 or deeper |
| Own Territory | Own 30 to midfield |
| Opp Territory | Midfield to opp 30 |
| Scoring Zone | Inside opp 20 |

Example: If they pass 70% on First down when backed up but only convert at 32%,
that is aggressive but ineffective  a defensive opportunity.
""")
            st.divider()

            st.subheader(" Explosive Play")
            st.markdown("Any play gaining 15 or more yards. If an offense gains 500 yards but 200 come from 3 explosive plays, they are big-play dependent  stop those and the offense stalls.")
            st.divider()

            st.subheader(" Personnel Group")
            st.markdown("""
Two-digit code: RBs + TEs on the field. Remaining skill players = WRs.

| Code | RBs | TEs | WRs | Common Name |
|-|-|-|-|-|
| 00 | 0 | 0 | 5 | Empty |
| 10 | 1 | 0 | 4 | Trips/Quads |
| 11 | 1 | 1 | 3 | Standard Spread |
| 12 | 1 | 2 | 2 | Pro Set |
| 13 | 1 | 3 | 1 | Double Y |
| 20 | 2 | 0 | 3 | Wing |
| 21 | 2 | 1 | 2 | Power Spread |
| 22 | 2 | 2 | 1 | Heavy/I-Form |
""")

        #  TAB 1: PERSONNEL 
        with tabs[1]:
            st.header(" Personnel Identity")
            st.subheader("Overall Usage")
            st.dataframe(pers_counts, use_container_width=False)
            st.divider()
            c1, c2 = st.columns(2)
            with c1:
                st.subheader(" Top 5 Run Personnel")
                run_pers = (
                    p_data[p_data[cols['type']] == 'RUN']
                    .groupby('PERSONNEL')
                    .agg(Plays=('PERSONNEL','count'), Avg_Gain=(cols['gain'],'mean'))
                    .sort_values('Plays', ascending=False).head(5)
                )
                run_pers['Avg_Gain'] = run_pers['Avg_Gain'].round(1)
                run_pers['Run %'] = (run_pers['Plays'] / run_pers['Plays'].sum() * 100).round(0).astype(int)
                st.dataframe(run_pers.style.background_gradient(cmap='RdYlGn', subset=['Avg_Gain']), use_container_width=False)
            with c2:
                st.subheader(" Top 5 Pass Personnel")
                pass_pers = (
                    p_data[p_data[cols['type']] == 'PASS']
                    .groupby('PERSONNEL')
                    .agg(Plays=('PERSONNEL','count'), Avg_Gain=(cols['gain'],'mean'))
                    .sort_values('Plays', ascending=False).head(5)
                )
                pass_pers['Avg_Gain'] = pass_pers['Avg_Gain'].round(1)
                pass_pers['Pass %'] = (pass_pers['Plays'] / pass_pers['Plays'].sum() * 100).round(0).astype(int)
                st.dataframe(pass_pers.style.background_gradient(cmap='RdYlGn', subset=['Avg_Gain']), use_container_width=False)
            st.divider()
            st.subheader("Run/Pass Tendency by Personnel")
            rp_split = (
                p_data.groupby('PERSONNEL')[cols['type']]
                .value_counts(normalize=True).unstack().fillna(0).mul(100).round(0).astype(int)
            )
            st.dataframe(rp_split.style.background_gradient(cmap='RdYlGn_r'), use_container_width=False)

        #  TAB 2: 3RD DOWN 
        with tabs[2]:
            st.header(" Third Down Efficiency")
            if not t3.empty:
                st.metric("Third Down Conversion Rate", f"{round(t3['Is_FD'].mean()*100)}pct")
                c1, c2 = st.columns(2)
                with c1:
                    st.table(t3_summary)
                with c2:
                    t3_tend = t3.groupby('Sit')[cols['type']].value_counts(normalize=True).unstack().fillna(0).mul(100).round(0).astype(int)
                    st.dataframe(t3_tend.style.background_gradient(cmap='RdYlGn_r').format("{:d}pct"), use_container_width=False)
                for sit in ["Third & Short (1-3)", "Third & Mid (4-7)", "Third & Long (7+)"]:
                    with st.expander(f"Top Calls: {sit}"):
                        st.table(t3[t3['Sit'] == sit][cols['play']].value_counts().head(3))
            else:
                st.info("No Third down plays found.")

        #  TAB 3: CHAIN MOVING 
        with tabs[3]:
            st.header(" Chain Moving (Frequency)")
            m1, m2 = st.columns(2)
            m1.metric("Overall FD Rate",      f"{round(p_data['Is_FD'].mean()*100)}pct")
            m2.metric("Overall Success Rate", f"{round(p_data['Is_Succ'].mean()*100)}pct")
            st.divider()
            st.dataframe(
                chain.style.background_gradient(cmap='RdYlGn', subset=['FD Rate %','Success Rate %']),
                use_container_width=False
            )

        #  TAB 4: RED/GREEN ZONE 
        with tabs[4]:
            st.header(" Red/Green Zone")
            rz = p_data[p_data[cols['field']].between(11, 20)].copy()
            gz = p_data[p_data[cols['field']].between(1, 10)].copy()
            c1, c2 = st.columns(2)
            with c1:
                st.subheader(" Red Zone (11-20)")
                if not rz.empty:
                    st.metric("TD/FD Rate", f"{round(rz['Is_FD'].mean()*100)}pct")
                    st.table(rz[cols['play']].value_counts().head(5).to_frame("Plays"))
                else:
                    st.info("No red zone plays.")
            with c2:
                st.subheader(" Green Zone (1-10)")
                if not gz.empty:
                    st.metric("Success Rate", f"{round(gz['Is_Succ'].mean()*100)}pct")
                    st.table(gz[cols['play']].value_counts().head(5).to_frame("Plays"))
                else:
                    st.info("No green zone plays.")

        #  TAB 5: WINNING PROBABILITY 
        with tabs[5]:
            st.header(" Winning Probability (AI)")

            st.subheader(" AI Scouting Intelligence")
            st.caption("Auto-detected behavioral patterns and tendencies from play-by-play data.")
            if not intel_df.empty:
                st.dataframe(
                    intel_df.set_index('Category'),
                    use_container_width=True,
                    column_config={
                        "Signal":        st.column_config.TextColumn("Signal",        width=200),
                        "Stat":          st.column_config.TextColumn("Stat",          width=160),
                        "Coaching Note": st.column_config.TextColumn("Coaching Note", width=500),
                    }
                )
            else:
                st.info("No intelligence signals detected yet.")

            st.divider()
            st.subheader(" Sequence Stress Score (SSS)")
            st.caption("What play types and formations are creating Third & long situations.")
            if not sss_summary.empty:
                c1, c2 = st.columns(2)
                with c1:
                    st.write("By Play Type")
                    st.dataframe(sss_summary.style.background_gradient(cmap='RdYlGn_r', subset=['Stress %']), use_container_width=False)
                with c2:
                    st.write("Top Formations Creating Stress")
                    st.dataframe(sss_by_form.style.background_gradient(cmap='RdYlGn_r', subset=['Stress_Count']), use_container_width=False)
                with st.expander(" Full Stress Play Log"):
                    sss_display = sss_df.reset_index(drop=True).astype(str)
                    st.dataframe(sss_display, use_container_width=False)
            else:
                st.info("No Third & long stress situations found.")

            st.divider()
            st.subheader(" Field Position Aggression Rating (FPAR)")
            st.caption("Pass rate, success rate, and avg gain by field zone and down.")
            if not fpar_df.empty:
                st.dataframe(fpar_df.style.background_gradient(cmap='RdYlGn', subset=['Success_Rate']), use_container_width=False)
                st.write("First Down Pass Rate by Zone")
                zone_First = fpar_df.reset_index()
                zone_First = zone_First[zone_First[cols['dn']] == 1][['Field_Zone','Pass_Rate','Success_Rate','Avg_Gain']]
                if not zone_First.empty:
                    st.dataframe(zone_First.set_index('Field_Zone').style.background_gradient(cmap='RdYlGn', subset=['Success_Rate']), use_container_width=False)
            else:
                st.info("Not enough data for field position analysis.")

        #  TAB 6: PIVOT LAB 
        with tabs[6]:
            st.header(" Pivot Lab")
            st.caption("Build your own custom views. Filter, group, and export any combination of data.")
            #  FILTERS 
            st.subheader(" Filters")
            fc1, fc2, fc3, fc4 = st.columns(4)
            with fc1:
                play_filter = st.radio("Play Type", ["ALL", "RUN", "PASS"], horizontal=True)
            with fc2:
                down_filter = st.multiselect("Down", [1, 2, 3, 4], default=[1, 2, 3, 4])
            with fc3:
                dist_filter = st.radio("Distance", ["ALL", "Short (1-3)", "Med (4-7)", "Long (8+)"], horizontal=False)
            with fc4:
                zone_filter = st.radio("Field Zone", ["ALL", "Own Territory", "Midfield", "Scoring Zone"], horizontal=False)

            #  APPLY FILTERS 
            view = p_data.copy()
            if play_filter != "ALL":
                view = view[view[cols['type']] == play_filter]
            if down_filter:
                view = view[view[cols['dn']].isin(down_filter)]
            if dist_filter == "Short (1-3)":
                view = view[view[cols['dist']] <= 3]
            elif dist_filter == "Med (4-7)":
                view = view[view[cols['dist']].between(4, 7)]
            elif dist_filter == "Long (8+)":
                view = view[view[cols['dist']] >= 8]
            if zone_filter == "Own Territory":
                view = view[view[cols['field']] <= -30]
            elif zone_filter == "Midfield":
                view = view[view[cols['field']].between(-29, 0)]
            elif zone_filter == "Scoring Zone":
                view = view[view[cols['field']] >= 1]

            #  SUMMARY BAR 
            if not view.empty:
                sm1, sm2, sm3, sm4, sm5 = st.columns(5)
                sm1.metric("Plays",        len(view))
                sm2.metric("Avg Gain",     f"{view[cols['gain']].mean():.1f} yds")
                sm3.metric("FD Rate",      f"{round(view['Is_FD'].mean()*100)}pct")
                sm4.metric("Success Rate", f"{round(view['Is_Succ'].mean()*100)}pct")
                sm5.metric("Explosive %",  f"{round(view['Is_Explosive'].mean()*100)}pct")
            else:
                st.warning("No plays match the selected filters.")

            st.divider()

            #  PIVOT BUILDER 
            st.subheader(" Build Your Table")
            pc1, pc2, pc3 = st.columns(3)
            with pc1:
                row_by = st.selectbox("Group by (rows)", [
                    cols['form'], 'PERSONNEL', cols['play'],
                    cols['dn'], cols['type'], cols['p_dir'], cols['hash']
                ])
            with pc2:
                break_by = st.selectbox("Break down by (columns)", [
                    "None", cols['type'], cols['dn'], "Distance Bucket", cols['hash'], cols['p_dir']
                ])
            with pc3:
                show_val = st.selectbox("Show me", [
                    "Avg Gain", "FD Rate %", "Success Rate %",
                    "Explosive Rate %", "Play Count"
                ])

            #  BUILD PIVOT 
            if not view.empty and row_by in view.columns:
                piv_view = view.copy()

                if break_by == "Distance Bucket":
                    piv_view['Distance Bucket'] = piv_view[cols['dist']].apply(
                        lambda x: "Short (1-3)" if x <= 3 else ("Med (4-7)" if x <= 7 else "Long (8+)")
                    )
                    break_col = "Distance Bucket"
                elif break_by == "None":
                    break_col = None
                else:
                    break_col = break_by

                val_map = {
                    "Avg Gain":         cols['gain'],
                    "FD Rate %":        'Is_FD',
                    "Success Rate %":   'Is_Succ',
                    "Explosive Rate %": 'Is_Explosive',
                    "Play Count":       cols['gain'],
                }
                agg_map = {
                    "Avg Gain":         'mean',
                    "FD Rate %":        'mean',
                    "Success Rate %":   'mean',
                    "Explosive Rate %": 'mean',
                    "Play Count":       'count',
                }

                val_col = val_map[show_val]
                agg_fn  = agg_map[show_val]

                if break_col and break_col in piv_view.columns:
                    pivot_result = piv_view.pivot_table(
                        index=row_by,
                        columns=break_col,
                        values=val_col,
                        aggfunc=agg_fn
                    )
                    if show_val != "Play Count":
                        pivot_result = (pivot_result * (100 if "Rate" in show_val else 1)).round(1)
                else:
                    if agg_fn == 'mean':
                        pivot_result = piv_view.groupby(row_by)[val_col].mean()
                        if "Rate" in show_val:
                            pivot_result = (pivot_result * 100).round(1)
                        else:
                            pivot_result = pivot_result.round(1)
                    else:
                        pivot_result = piv_view.groupby(row_by)[val_col].count()
                    pivot_result = pivot_result.to_frame(show_val)

                pivot_result = pivot_result.reset_index()
                numeric_cols = pivot_result.select_dtypes('number').columns.tolist()

                st.dataframe(
                    pivot_result.style.background_gradient(cmap='RdYlGn', subset=numeric_cols),
                    use_container_width=True
                )

                #  EXPORT THIS VIEW 
                st.divider()
                pivot_export = BytesIO()
                with pd.ExcelWriter(pivot_export, engine='openpyxl') as writer:
                    pivot_result.to_excel(writer, sheet_name='Pivot View', index=False)
                    view.reset_index(drop=True).to_excel(writer, sheet_name='Filtered Plays', index=False)
                st.download_button(
                    label=" Export This View to Excel",
                    data=pivot_export.getvalue(),
                    file_name="FormationIQ_PivotView.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                st.info("Select valid grouping options above to build your table.")

#  TAB 7: Vic Fangio ASSISTANT 
    with tabs[7]:
             st.header(" Vic Fangio Assistant")
             st.caption("Ask questions about this opponent. Answers are from a coach modeled after Vic Fangio.")
    if "fangio_chat" not in st.session_state:
        st.session_state.fangio_chat = []

    current_view = view if "view" in locals() else p_data

    for msg in st.session_state.fangio_chat:
        with st.chat_message(msg["role"]):
            st.markdown(msg["content"])

    question = st.chat_input("Ask about tendencies, game plan, or how to attack this defense...")

    if question:
        st.session_state.fangio_chat.append({"role": "user", "content": question})
        with st.spinner("Coach Fangio is reviewing the film..."):
            summary_text = summarize_view(current_view)
            answer = call_fangio_llm(summary_text, question)
        st.session_state.fangio_chat.append({"role": "assistant", "content": answer})
        with st.chat_message("assistant"):
            st.markdown(answer)

        #  TAB 8: DRIVE LEVERAGE 
        with tabs[8]:
            st.header(" Drive Leverage Score (DLS)")
            st.subheader("Per-Drive Summary")
            st.dataframe(drive_dla.reset_index().astype({c: str for c in drive_dla.reset_index().select_dtypes('object').columns}), use_container_width=False)
            st.divider()
            st.subheader("Personnel Leverage Profile")
            st.dataframe(pers_dla.sort_values('DLS', ascending=False).reset_index().astype({c: str for c in pers_dla.sort_values('DLS', ascending=False).reset_index().select_dtypes('object').columns}), use_container_width=False)
            st.divider()
            with st.expander(" Personnel + Formation Leverage (min 5 plays)"):
                pf_display = pf_dla.sort_values('DLS', ascending=False).reset_index()
                pf_display = pf_display.astype({c: str for c in pf_display.select_dtypes('object').columns})
                for col in ['DLS','Avg_Gain','FD_Rate','Success_Rate','Explosive_Rt','High_Lev%','Low_Lev%']:
                    if col in pf_display.columns:
                        pf_display[col] = pd.to_numeric(pf_display[col], errors='coerce')
                st.dataframe(pf_display.style.background_gradient(cmap='RdYlGn', subset=['DLS']), use_container_width=False)


    else:
        st.info("Upload a Hudl CSV or Excel file to get started.")
